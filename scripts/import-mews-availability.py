#!/usr/bin/env python3
"""
import-mews-availability.py
===========================
Reads a Mews availability report Excel file and imports the data into
the hotel-platform PostgreSQL database.

Usage:
  python3 import-mews-availability.py path/to/Availability_report.xlsx [--dry-run]

Environment variables (or edit defaults below):
  DB_HOST      PostgreSQL host          (default: localhost)
  DB_PORT      PostgreSQL port          (default: 5432)
  DB_NAME      Database name            (default: hotel_platform)
  DB_USER      Database user            (default: postgres)
  DB_PASS      Database password        (default: postgres)

Run on the server:
  docker exec hotel-staging-api bash -c "
    apt-get install -y python3-pip 2>/dev/null;
    pip3 install openpyxl psycopg2-binary --break-system-packages;
    python3 /opt/hotel-platform-staging/scripts/import-mews-availability.py \
      /opt/hotel-platform-staging/data/Availability_report_2026.xlsx
  "
"""

import os
import sys
import argparse
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    os.system("pip3 install openpyxl --break-system-packages -q")
    import openpyxl

try:
    import psycopg2
    from psycopg2.extras import execute_values
except ImportError:
    print("Installing psycopg2-binary...")
    os.system("pip3 install psycopg2-binary --break-system-packages -q")
    import psycopg2
    from psycopg2.extras import execute_values

# ---------------------------------------------------------------------------
# DB connection defaults (override via env vars)
# ---------------------------------------------------------------------------
DB_CONFIG = {
    'host':     os.environ.get('DB_HOST', 'localhost'),
    'port':     int(os.environ.get('DB_PORT', '5432')),
    'dbname':   os.environ.get('DB_NAME', 'hotel_platform'),
    'user':     os.environ.get('DB_USER', 'postgres'),
    'password': os.environ.get('DB_PASS', 'postgres'),
}

# ---------------------------------------------------------------------------
# Mews category → display name mapping
# Edit this to match your Resources table names
# ---------------------------------------------------------------------------
CATEGORY_LABELS = {
    'BUD':            'Budget Room',
    'STD':            'Standard Room',
    'JS+':            'Junior Suite Plus',
    'ES+':            'Executive Suite Plus',
    'ES':             'Executive Suite',
    'JS':             'Junior Suite',
    'STD+':           'Standard Plus',
    'STD3':           'Standard Triple',
    'DLX':            'Deluxe Room',
    'DLX+':           'Deluxe Plus',
    'SESU24':         'Sea Suite 2024',
    'SESU25':         'Sea Suite 2025',
    'Senior Suite 1': 'Senior Suite 1',
}


def parse_excel(path: str) -> dict:
    """
    Returns {category_code: {date_str: available_count}}
    Reads the 'Availability' sheet from the Mews report.
    """
    print(f"Reading Excel: {path}")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    if 'Availability' not in wb.sheetnames:
        raise ValueError(f"Sheet 'Availability' not found. Sheets: {wb.sheetnames}")

    ws = wb['Availability']
    rows = list(ws.iter_rows(values_only=True))

    header_row = rows[0]  # ('Service', 'Space category', date1, date2, ...)
    dates = []
    for cell in header_row[2:]:
        if cell is None or str(cell) == 'Total':
            break
        if isinstance(cell, datetime):
            dates.append(cell.strftime('%Y-%m-%d'))
        else:
            break

    print(f"  Found {len(dates)} dates: {dates[0]} → {dates[-1]}")

    result = {}
    for row in rows[1:]:
        category = row[1]
        if category is None or str(category).strip() == '':
            continue  # skip Total row

        category = str(category).strip()
        avail = {}
        for i, date_str in enumerate(dates):
            val = row[i + 2]
            avail[date_str] = int(val) if val is not None else 0

        result[category] = avail
        print(f"  Category '{category}': {len(avail)} dates loaded")

    wb.close()
    return result


def get_or_create_resources(conn, data: dict) -> dict:
    """
    Returns {category_code: resource_id}.
    Matches resources by mews_category column if it exists,
    otherwise falls back to matching by name containing the category code.
    Creates resources that don't exist yet.
    """
    cur = conn.cursor()

    # Check if mews_category column exists
    cur.execute("""
        SELECT column_name FROM information_schema.columns
        WHERE table_name = 'resources' AND column_name = 'mews_category'
    """)
    has_mews_col = cur.fetchone() is not None

    if has_mews_col:
        print("\nUsing mews_category column for resource matching")
    else:
        print("\nNo mews_category column found — matching by name")
        print("  TIP: Add 'mews_category VARCHAR(50)' to resources table for better matching")

    # Get all existing resources
    cur.execute("SELECT id, name, mews_category FROM resources" if has_mews_col
                else "SELECT id, name FROM resources")
    rows = cur.fetchall()

    mapping = {}  # category_code → resource_id

    for cat in data.keys():
        resource_id = None

        if has_mews_col:
            for row in rows:
                rid, rname, rmews = row
                if rmews and rmews.strip() == cat:
                    resource_id = rid
                    print(f"  ✓ '{cat}' → resource #{rid} '{rname}' (by mews_category)")
                    break
        else:
            # Fuzzy match by name containing category code
            cat_lower = cat.lower().replace('+', 'plus').replace(' ', '')
            for row in rows:
                rid, rname = row
                rname_lower = rname.lower().replace('+', 'plus').replace(' ', '').replace('-', '')
                if cat_lower in rname_lower or rname_lower in cat_lower:
                    resource_id = rid
                    print(f"  ✓ '{cat}' → resource #{rid} '{rname}' (by name match)")
                    break

        if resource_id is None:
            print(f"  ✗ '{cat}' → no matching resource found")
            print(f"    Run with --create-resources to auto-create missing resources")
            mapping[cat] = None
        else:
            mapping[cat] = resource_id

    cur.close()
    return mapping


def create_missing_resources(conn, data: dict, mapping: dict) -> dict:
    """Create resources for unmapped categories."""
    cur = conn.cursor()

    # Need resource_group_id — use first available group
    cur.execute("SELECT id FROM resource_groups LIMIT 1")
    row = cur.fetchone()
    if not row:
        print("  ERROR: No resource groups found. Create a resource group first.")
        cur.close()
        return mapping

    group_id = row[0]

    for cat, rid in list(mapping.items()):
        if rid is None:
            label = CATEGORY_LABELS.get(cat, cat)
            cur.execute("""
                INSERT INTO resources (name, mews_category, resource_group_id, is_active)
                VALUES (%s, %s, %s, true)
                ON CONFLICT DO NOTHING
                RETURNING id
            """, (label, cat, group_id))
            row = cur.fetchone()
            if row:
                mapping[cat] = row[0]
                print(f"  ✓ Created resource '{label}' (mews={cat}) → id #{row[0]}")
            else:
                print(f"  ✗ Could not create resource for '{cat}'")

    conn.commit()
    cur.close()
    return mapping


def import_availability(conn, data: dict, mapping: dict, dry_run: bool = False) -> None:
    """
    Upserts availability_days records.
    Schema assumed: (resource_id, date, available_count, updated_at)
    with UNIQUE constraint on (resource_id, date).
    """
    cur = conn.cursor()

    # Check table structure
    cur.execute("""
        SELECT column_name, data_type
        FROM information_schema.columns
        WHERE table_name = 'availability_days'
        ORDER BY ordinal_position
    """)
    columns = [(r[0], r[1]) for r in cur.fetchall()]
    col_names = [c[0] for c in columns]
    print(f"\navailability_days columns: {col_names}")

    if not columns:
        print("ERROR: availability_days table not found!")
        print("Check the entity file and make sure TypeORM has synced the table.")
        cur.close()
        return

    has_available_count = 'available_count' in col_names
    has_is_available = 'is_available' in col_names

    records = []
    skipped = 0

    for cat, avail in data.items():
        resource_id = mapping.get(cat)
        if resource_id is None:
            print(f"  Skipping '{cat}' (no matching resource)")
            skipped += len(avail)
            continue

        for date_str, count in avail.items():
            if has_available_count:
                records.append((resource_id, date_str, count, datetime.utcnow()))
            elif has_is_available:
                records.append((resource_id, date_str, count > 0, datetime.utcnow()))
            else:
                records.append((resource_id, date_str, datetime.utcnow()))

    total = len(records)
    print(f"\nPrepared {total} records ({skipped} skipped due to unmapped resources)")

    if dry_run:
        print("\n[DRY RUN] Would upsert the following sample (first 10):")
        for r in records[:10]:
            print(f"  {r}")
        print(f"  ... and {max(0, total - 10)} more")
        cur.close()
        return

    if not records:
        print("Nothing to import.")
        cur.close()
        return

    print(f"Importing {total} records...")

    if has_available_count:
        execute_values(cur, """
            INSERT INTO availability_days (resource_id, date, available_count, updated_at)
            VALUES %s
            ON CONFLICT (resource_id, date)
            DO UPDATE SET available_count = EXCLUDED.available_count,
                          updated_at = EXCLUDED.updated_at
        """, records)
    elif has_is_available:
        execute_values(cur, """
            INSERT INTO availability_days (resource_id, date, is_available, updated_at)
            VALUES %s
            ON CONFLICT (resource_id, date)
            DO UPDATE SET is_available = EXCLUDED.is_available,
                          updated_at = EXCLUDED.updated_at
        """, records)
    else:
        print("ERROR: Don't know how to insert into availability_days — unrecognised schema")
        print(f"Columns found: {col_names}")
        cur.close()
        return

    conn.commit()
    print(f"✓ Imported {total} availability records")
    cur.close()


def print_summary(data: dict, mapping: dict) -> None:
    print("\n=== MAPPING SUMMARY ===")
    for cat, avail in data.items():
        rid = mapping.get(cat)
        total = len(avail)
        sold_out = sum(1 for v in avail.values() if v == 0)
        print(f"  {cat:20s} → resource_id={rid or 'UNMAPPED':>4}  "
              f"({total} days, {sold_out} sold out)")


def main():
    parser = argparse.ArgumentParser(description='Import Mews availability report into hotel platform DB')
    parser.add_argument('excel_file', help='Path to Mews Availability_report.xlsx')
    parser.add_argument('--dry-run', action='store_true', help='Preview without writing to DB')
    parser.add_argument('--create-resources', action='store_true',
                        help='Auto-create resources for unmapped Mews categories')
    args = parser.parse_args()

    # Parse Excel
    data = parse_excel(args.excel_file)

    # Connect to DB
    print(f"\nConnecting to DB: {DB_CONFIG['host']}:{DB_CONFIG['port']}/{DB_CONFIG['dbname']}")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        print("  ✓ Connected")
    except Exception as e:
        print(f"  ✗ Connection failed: {e}")
        print("\nHint: If running inside Docker, try:")
        print("  docker exec hotel-staging-api python3 this_script.py")
        print("  Or set DB_HOST to the container's internal IP")
        sys.exit(1)

    # Map categories to resources
    mapping = get_or_create_resources(conn, data)

    if args.create_resources:
        mapping = create_missing_resources(conn, data, mapping)

    print_summary(data, mapping)

    # Import
    import_availability(conn, data, mapping, dry_run=args.dry_run)

    conn.close()
    print("\nDone.")


if __name__ == '__main__':
    main()
